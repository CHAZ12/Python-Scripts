# !/usr/bin/env python3
import os
import subprocess
import sys
import re
import time
from datetime import datetime, timedelta
import json

print('Getting or installing necessary modules')
#   Inorder for the script to run properly you need all modules below, so make sure they are installed
try:
    import requests

    print('module requests is installed')
except ModuleNotFoundError:
    subprocess.call([sys.executable, '-m', 'pip', 'install', 'requests'])

try:
    import pandas

    print('module pandas installed')
except ModuleNotFoundError:
    print('module pandas is not installed')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'pandas'])

try:
    from openpyxl import Workbook

    print('module openpyxl installed')
except ModuleNotFoundError:
    subprocess.call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])

try:
    import xlsxwriter

    print('module xlsxwriter installed')
except ModuleNotFoundError:
    subprocess.call([sys.executable, '-m', 'pip', 'install', 'xlsxwriter'])

try:
    from  datetime import datetime

    print('module datetime installed')
except ModuleNotFoundError:
    subprocess.call([sys.executable, '-m', 'pip', 'install', 'datetime'])

print("\n------ THIS IS A SCRIPT TO GET LIST OF BUG REPORTING FROM DISCORD -----")

def post_request_func(domain):
    print('(Fetching a valid url response state for: ' + domain + ")")
    homepage = requests.get(domain).text
    isLive = False
    if not homepage:
        print('(Page not found)\n')
    else:
        print('(Page found. Continuing...)\n')
        isLive = True
    if isLive:
        token = 'discord token'
        headers1 = { 
            'Content-Type': 'application/json', 
            'authorization':token
               }
        # URL Query
        query1 = 'query {}'
        data1 = json.dumps([{"query": query1}])
        global resp1
        resp1 = requests.get("http://discord.com/api/v9/channels//messages?limit=100", data=data1,headers=headers1)
        getUserData()
    
def getUserData():
    print("(GetUserData)")
    global setSecondlastUser
    setSecondlastUser = "NotSet"
    getLastUserFile()
    for i in range(len(resp1.json())):
        username(i)
        attachments(i)
        content(i)
        finalLabel()

def getLastUserFile():
    global newLastUser1, newLastUser2
    try:
        with open("LastUser.txt", "r+", encoding="utf-8") as file:
            lastUser = file.readlines()
            if lastUser:
                newLastUser1= re.sub('-(.*?T[^-\s]+)([.]+)(\d+)\+(.+)', r'-\1;\3;\4', lastUser[0].strip('\n'))
                newLastUser2= re.sub('-(.*?T[^-\s]+)([.]+)(\d+)\+(.+)', r'-\1;\3;\4', lastUser[1].strip('\n')) #Just incase someone deletes their post look for next closest user 
            else:
                newLastUser1 = newLastUser2 = "NONE"
    except FileNotFoundError:
        newLastUser1 =newLastUser2 = "NONE"
    
def username(i):
    global name, time
    time = str(resp1.json()[i]['timestamp'])
    name = str(resp1.json()[i]['author']['global_name'])
    if name == "None":# User did not set a globa name
        name = str(resp1.json()[i]['author']['username'])
    currentUser = "Username:" + name + "," + time
    # Have to replace the dot and plus with something else so I can match it"
    newCurrentUser= re.sub('-(.*?T[^-\s]+)([.]+)(\d+)\+(.+)', r'-\1;\3;\4', currentUser) 
    if(bool(re.search(newCurrentUser, newLastUser1)) or bool(re.search(newCurrentUser, newLastUser2))):
        print("(Found last user before reaching the end of total users.)")
        if i==0:
            print("First user was known as last user ")
            print("(Exiting... )")
            sys.exit()
    
    elif i == 0 or i==1: # check most recent user recorded
        if i==0:
            print("(LastUser:)" + newLastUser1)
            print("(Total Usernames:{0})".format(len(resp1.json())-1))
            global setlastUser
            setlastUser = newCurrentUser
        else:
            print("(SecondToLastUser:)" + newCurrentUser)
            setSecondlastUser = newCurrentUser
            with open("LastUser.txt", "w+", encoding="utf-8") as file:
                file.write(setlastUser + "\n" + setSecondlastUser)
            
    elif i == len(resp1.json())-1: # reached the max number of logs put a double line for cut off and write last user
        print("(Reached the end of the List... wrote last user)")
    print("(Got Username:{0})".format(i))
    
def writeComplete(newLogs):
    print("(writeComplete)")
    oldLogs = open("oldLogs.txt", "r", encoding="utf-8")
    for line in oldLogs:
        newLogs.write(line) 
    oldLogs.close()
    newLogs.close() 
    os.remove("oldLogs.txt") # remove old backup logs
    print("(sucessfully wrote and updated newLogs)")
    #with open("LastUser.txt", "w+", encoding="utf-8") as file:
        #file.write(setlastUser + "\n" + setSecondlastUser)
    print("(Exiting...)")
    sys.exit()  
                 
def content(i): # Look for discord message
    print("(content)")
    global message
    message = str(resp1.json()[i]['content']).strip("\n")
    message = message.lower().replace(";", "")
    if message == "":
        message = "[]"
    findContent(message,i)
    
def attachments(i): # Look for any attachments such as Images/Urls
    print("(Attachments)")
    global url
    url = str(resp1.json()[i]['attachments'])
    if url != '[]':
        url = str(resp1.json()[i]['attachments'][0]['url'])

def findContent(message, i): # Determine what category the content should be in
    print('(findContent)')
    catFound = False
    notmodded = False
    global catarray
    catarray = []
    groups =[
        "not modded|no mods|tried mods|unmodded", #0
        "modded|bepinex|mod|mods|r2modmanger|thunderstore", #1
        "disconnects|disconnect|disconnects|dc|disconnected|dc|crashing|join|lobbies|hosts|host|crash|exit the game|connection|freezing|crash|LAN|lan|server|corrupted|could not be loaded|servers|tabbing out|client side|version|gifted",
        "UI|framerate|fps|performance|voice|drivers|black screen|lag|audio", #3
        "rebind|keybind|keybinds|scroll wheel|controller|bind to|microphone",#4
        "nutcracker|mimic|mask|masked|bracken|turrets|turret|bees|dog|crawler|baboons hawk|spider|coil head|eyeless dog|forest keeper|girl|hoarding bug|jester|snare flea|spore lizard|thumper|cobweb|stomper|monster|giants|blind dog",#5
        "locker|inverse|teleporter|ship|scan|price|teleported|teleporting|televison",#6
        "zapgun|flashlight|flash-light|spraypaint|spray paint|ladder|items|presents|shutgun|keys|gift|jetpack|shovel|paint|shotguns", #7
        "duped|dupping|duplicate|a lot of items|exit the game|items|item|outside of the map|super fast|dupes|glitch",#8
        "rank|exp|title|level"#9
    ]
    
    #Group each message 
    for j in range(len(groups)):
        cat = re.findall(r"(?:$|^| )" + groups[j] +"(?:$|^| |,)", message)  # Replace with category
        if cat:        
            # Find the category
            catFound = True
            match j:
                case 0:# Unmodded
                    notmodded = True
                case 1:
                    catarray.append('Modded')
                    if notmodded:
                        catarray.remove("Modded")
                        pass
                case 2:
                    catarray.append('Server')
                case 3:
                    catarray.append('Graphics')
                case 4:
                    catarray.append('Controls')
                case 5:
                    catarray.append('Creatures')
                case 6:
                    catarray.append('ShipItem')
                case 7:
                    catarray.append('Equipment')
                case 8:
                    catarray.append('Glitches')
                case 9:
                    catarray.append('Rank')     
    if not catFound:
        print("(Setting catFound to MISC))")
        catarray.append("Misc")
        pass
    else:
        print("(Checking Modded)")
        if(int(len(catarray) != 0)):
            for j in range(len(catarray)):
                if catarray[j] == "Modded" and notmodded != True:
                    catarray[0] = "Modded"
                    pass
        else:
            catarray.append("MISC")
    print("Found Labels:" + str(catarray))

def finalLabel():
    print("Time:{0}\nMain Label:{1}\nUsername:{2}\nMessage:{3}\nAttachments:{4}\nFound Labels:{5}\n".format(time, catarray[0],name,message,url,catarray))
    Start_Excel_Data()

def Start_Excel_Data():
    print("Start Getting Excel Data")
    try:
        open("newLogs.txt", "r", encoding="utf-8")
        Data_file = "newLogs.txt"
    except:
        print("Could not find newLogs..Trying oldLogs")
        open("oldLogs.txt", "r", encoding="utf-8")
        Data_file = "oldLogs.txt"
    else:
        print("NO FILES FOUND EXITING..")
        sys.exit()

    with open(Data_file, "r+", encoding="utf-8") as file:
        data = file.readlines()
        #print(data)
        for line in data: 
            try:
                    rName = re.search("(?<=Username:)(.+)(?=;;Message)", line, re.MULTILINE)[0]
                    rMainLabel = re.search("(?<=Main Label:)(.*)(?=;;User)", line, re.MULTILINE)[0]
                    rMessage = re.search('(?<=Message:)([^;;]*)(?=;;)', line.strip(), re.MULTILINE)[0]
                    rFoundLabels = re.search("(?<=Found Labels:)(.+)", line, re.MULTILINE)[0].strip("\s")
                    rDate = re.search("(?<=Time:)(\S+)(?=;;)", line, re.MULTILINE)[0]
                    Convert_Time(rDate)
                    rAttachments = re.search("(?<=Attachments:)(.+)(?=;;)", line, re.MULTILINE)[0]
                    create_excel_func(rName, rMainLabel, rMessage, rFoundLabels, rAttachments)
            except:
                    pass

import xlsxwriter
from openpyxl import Workbook
def create_excel_func(Name, MLabel, Msg, FLabels, Attachments):
    #print(Name)
    #   GET READY TO CREATE OR OPEN A EXCEL FILE TO POST NEW INFORMATION
    #print('Checking Excel data')
    nameExcel = "LethalCompanyTestStats.xlsx"
    try:
        open(nameExcel)
        write_excel_file(Name,MLabel,Msg,FLabels,Attachments)
    except IOError:
        #print('Created new excel file')
        workbook1 = xlsxwriter.Workbook('LethalCompanyTestStats.xlsx')
        worksheet1 = workbook1.add_worksheet("ALL DATA")
        # Add a table to the worksheet.
        worksheet1.add_table(
            "A1:F900",
            {
                "columns": [
                    {"header": "Name"},
                    {"header": "Main Label"},
                    {"header": "Message"},
                    {"header": "Found Labels"},
                    {"header": "Date"},
                    {"header": "Attachments"},
                ],
            },
        )
        workbook1.close()
        print("Create new sheet")
        wb = openpyxl.load_workbook('LethalCompanyTestStats.xlsx')
        wb.create_sheet("Search")
        wb.save('LethalCompanyTestStats.xlsx')
        wb.close

import openpyxl
def write_excel_file(Name,MLabel,Msg,FLabels,Attachments):
    print("Begin Appending Excel")
    #Append new data to old file
    wb = openpyxl.load_workbook('LethalCompanyTestStats.xlsx')
    sheet = wb.active
    new_row = (Name, MLabel, Msg, FLabels, formatted_time, Attachments)
    sheet.append(new_row)
    wb.save('LethalCompanyTestStats.xlsx')
    
def Convert_Time(rDate):
    # Parse the input string into a datetime object
    dt_object = datetime.strptime(rDate, "%Y-%m-%dT%H:%M:%S.%f%z")
    # Convert to local time (adjusting for the UTC offset)
    dt_local = dt_object - timedelta(hours=dt_object.utcoffset().seconds // 3600)
    # Format the datetime object as desired
    global formatted_time
    formatted_time = dt_local.strftime('%m/%d/%y %I:%M %p')

post_request_func('http://discord.com/api/v9/channels/774242154561798144/messages?limit=1')

#(?<=Time:)(.+)|(?<=Main Label)(.+)|(?<=Username:)(.+)\n((?:.*\n)*?)Attachments:|(?<=Attachments:\s)(.+)|(?<=Found Labels:\[)('.+)(?=\])
#([\s\S]*?(?<=Username~).+[\s\S])(.+)\s(?=Attachments) Add message to text
